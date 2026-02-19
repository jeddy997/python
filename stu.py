import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

data = [
    ["A", 78, 65, 80],
    ["B", 56, 70, 75],
    ["C", 90, 88, 92],
    ["D", 67, 72, 60],
    ["E", 85, 79, 88],
    ["F", 73, 68, 70],
    ["G", 60, 75, 65],
    ["H", 88, 84, 91],
    ["I", 92, 90, 94],
    ["J", 70, 66, 72]
]

columns = ["Student", "Math", "English", "ICT"]

def create_excel_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(columns)

    for row in data:
        ws.append(row)

    wb.save("student_scores.xlsx")
    print("Excel file created successfully.")

def process_data():
    df = pd.read_excel("student_scores.xlsx")


    df["Total"] = df[["Math", "English", "ICT"]].sum(axis=1)

    
    df["Average"] = df["Total"] / 3

    subject_averages = df[["Math", "English", "ICT"]].mean()

    def assign_grade(avg):
        if avg >= 80:
            return "A"
        elif avg >= 70:
            return "B"
        elif avg >= 60:
            return "C"
        else:
            return "D"

    df["Grade"] = df["Average"].apply(assign_grade)

    print("\nProcessed Data:")
    print(df)

    print("\nSubject Averages:")
    print(subject_averages)

    return df

def perform_statistical_analysis(df):
    correlation = df[["Math", "English", "ICT"]].corr()

    print("\nCorrelation Matrix:")
    print(correlation)

   
    corr_pairs = correlation.unstack()
    corr_pairs = corr_pairs[corr_pairs < 1]  # remove 1.0 self correlation
    highest_corr = corr_pairs.idxmax()

    print("\nHighest Correlation Between:", highest_corr)

    return correlation

def generate_visualizations(df):
    subject_averages = df[["Math", "English", "ICT"]].mean()

    
    plt.figure()
    subject_averages.plot(kind="bar")
    plt.title("Subject Averages")
    plt.ylabel("Average Marks")
    plt.show()

    
    plt.figure()
    plt.plot(df["Student"], df["Math"])
    plt.plot(df["Student"], df["English"])
    plt.plot(df["Student"], df["ICT"])
    plt.title("Student Performance Comparison")
    plt.xlabel("Students")
    plt.ylabel("Marks")
    plt.legend(["Math", "English", "ICT"])
    plt.show()

    
    plt.figure()
    df["Grade"].value_counts().plot(kind="pie", autopct="%1.1f%%")
    plt.title("Grade Distribution")
    plt.ylabel("")
    plt.show()



def export_summary(df, correlation):
    wb = load_workbook("student_scores.xlsx")
    ws = wb.create_sheet("Summary")

    
    ws.append(["Subject Averages"])
    subject_averages = df[["Math", "English", "ICT"]].mean()

    for subject, avg in subject_averages.items():
        ws.append([subject, avg])

    
    top_student = df.loc[df["Average"].idxmax()]
    ws.append([])
    ws.append(["Top Performing Student"])
    ws.append([top_student["Student"], top_student["Average"]])

    
    ws.append([])
    ws.append(["Correlation Matrix"])
    ws.append(["", "Math", "English", "ICT"])

    for index, row in correlation.iterrows():
        ws.append([index] + list(row))

    wb.save("student_scores.xlsx")
    print("\nSummary sheet added successfully.")



create_excel_file()
df = process_data()
correlation = perform_statistical_analysis(df)
generate_visualizations(df)
export_summary(df, correlation)
